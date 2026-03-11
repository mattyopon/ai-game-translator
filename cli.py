#!/usr/bin/env python3
"""AI Game Translation Tool — ゲーム翻訳AIツール

Usage:
    python cli.py translate input.xlsx [-o output.xlsx] [--category skill_effect]
    python cli.py learn corrected.xlsx [--source-col A --translation-col B]
    python cli.py glossary add "必殺技" "Special Move"
    python cli.py glossary list
    python cli.py glossary import glossary.csv
    python cli.py stats
    python cli.py interactive
"""

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl
import yaml
from tqdm import tqdm

from glossary import Glossary
from translator import GameTranslator

# ---------------------------------------------------------------------------
# Paths & configuration helpers
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CONFIG_PATH = BASE_DIR / "config.yaml"
MEMORY_PATH = DATA_DIR / "translation_memory.json"
GLOSSARY_PATH = DATA_DIR / "glossary.json"


def load_config() -> dict:
    """Load configuration from config.yaml with environment variable overrides."""
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
    else:
        cfg = {}

    # Environment variable overrides
    if os.environ.get("ANTHROPIC_API_KEY"):
        cfg.setdefault("api", {})["key"] = os.environ["ANTHROPIC_API_KEY"]

    return cfg


def load_memory() -> dict:
    """Load translation memory from JSON."""
    if MEMORY_PATH.exists():
        with open(MEMORY_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"version": 1, "entries": []}


def save_memory(memory: dict) -> None:
    """Persist translation memory to JSON."""
    MEMORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(MEMORY_PATH, "w", encoding="utf-8") as f:
        json.dump(memory, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# Model shorthand map (for --model flag)
# ---------------------------------------------------------------------------

MODEL_MAP = {
    "opus": "claude-opus-4-20250514",
    "sonnet": "claude-sonnet-4-20250514",
    "haiku": "claude-haiku-4-20250514",
}


def _create_translator(model_override: str | None = None) -> GameTranslator:
    """Create a GameTranslator, optionally overriding the model."""
    translator = GameTranslator(config_path=str(CONFIG_PATH))
    if model_override and model_override in MODEL_MAP:
        translator._model = MODEL_MAP[model_override]
    return translator


# ---------------------------------------------------------------------------
# Auto-detect source column
# ---------------------------------------------------------------------------

def detect_source_column(ws) -> int:
    """Heuristic: find the first column whose header or first data cell contains Japanese."""
    import re

    jp_pattern = re.compile(r"[\u3000-\u9fff\uf900-\ufaff]")
    for col_idx in range(1, ws.max_column + 1):
        for row_idx in range(1, min(ws.max_row + 1, 6)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and jp_pattern.search(str(val)):
                return col_idx
    return 1  # fallback


def col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter (A, B, ..., Z, AA, ...) to 1-based index."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


# ---------------------------------------------------------------------------
# Subcommands
# ---------------------------------------------------------------------------

def cmd_translate(args: argparse.Namespace) -> None:
    """Translate an Excel file."""
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path.with_name(
        input_path.stem + "_translated" + input_path.suffix
    )

    wb = openpyxl.load_workbook(str(input_path))
    ws = wb.active

    # Determine source column
    if args.source_col:
        src_col = col_letter_to_index(args.source_col)
    else:
        src_col = detect_source_column(ws)

    # Translation output column = next empty column after last
    out_col = ws.max_column + 1
    conf_col = out_col + 1
    ws.cell(row=1, column=out_col, value="Translation (EN)")
    ws.cell(row=1, column=conf_col, value="Confidence")

    rows_to_translate: list[tuple[int, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=src_col).value
        if val and str(val).strip():
            rows_to_translate.append((row_idx, str(val).strip()))

    if not rows_to_translate:
        print("No translatable text found.")
        return

    if args.dry_run:
        print(f"Dry run — would translate {len(rows_to_translate)} rows")
        for row_idx, text in rows_to_translate[:10]:
            preview = text[:60] + ("..." if len(text) > 60 else "")
            print(f"  Row {row_idx}: {preview}")
        if len(rows_to_translate) > 10:
            print(f"  ... and {len(rows_to_translate) - 10} more rows")
        return

    # Create translator (uses config.yaml, supports both API and CLI backends)
    translator = _create_translator(model_override=args.model)

    translated_count = 0
    memory_hits = 0

    for row_idx, text in tqdm(rows_to_translate, desc="Translating", unit="row"):
        try:
            result = translator.translate(
                source=text,
                category=args.category or "",
            )
            translation = result["translation"]
            confidence = result["confidence"]

            ws.cell(row=row_idx, column=out_col, value=translation)
            ws.cell(row=row_idx, column=conf_col, value=round(confidence, 2))

            if confidence >= 1.0:
                memory_hits += 1
            else:
                translated_count += 1
                # Auto-save new translations to memory
                translator.memory.add(
                    source=text,
                    target=translation,
                    category=args.category or "",
                )
        except Exception as e:
            print(f"\n  Error translating row {row_idx}: {e}", file=sys.stderr)
            ws.cell(row=row_idx, column=out_col, value=f"[ERROR] {e}")

    wb.save(str(output_path))

    print(f"\nDone! {translated_count} translated, {memory_hits} from memory.")
    print(f"Output saved to: {output_path}")


def cmd_learn(args: argparse.Namespace) -> None:
    """Import approved/corrected translations into memory."""
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(input_path))
    ws = wb.active

    src_col = col_letter_to_index(args.source_col) if args.source_col else 1
    tl_col = col_letter_to_index(args.translation_col) if args.translation_col else 2
    corrected_col = col_letter_to_index(args.corrected_col) if args.corrected_col else None

    memory = load_memory()
    existing_sources = {e["source"] for e in memory.get("entries", [])}

    imported = 0
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        source = ws.cell(row=row_idx, column=src_col).value
        if not source or not str(source).strip():
            continue
        source = str(source).strip()

        # Prefer corrected column if available, otherwise use translation column
        if corrected_col:
            translation = ws.cell(row=row_idx, column=corrected_col).value
            if not translation or not str(translation).strip():
                translation = ws.cell(row=row_idx, column=tl_col).value
        else:
            translation = ws.cell(row=row_idx, column=tl_col).value

        if not translation or not str(translation).strip():
            continue
        translation = str(translation).strip()

        if source in existing_sources:
            # Update existing entry
            for entry in memory["entries"]:
                if entry["source"] == source:
                    entry["translation"] = translation
                    entry["category"] = args.category or entry.get("category", "")
                    entry["auto_approved"] = False
                    break
            updated += 1
        else:
            memory["entries"].append({
                "source": source,
                "translation": translation,
                "category": args.category or "",
                "auto_approved": False,
            })
            existing_sources.add(source)
            imported += 1

    save_memory(memory)
    print(f"Imported {imported} new entries, updated {updated} existing entries.")


def cmd_glossary(args: argparse.Namespace) -> None:
    """Manage glossary entries."""
    glossary = Glossary(str(GLOSSARY_PATH))

    if args.glossary_action == "add":
        glossary.add(args.source, args.target, notes=args.notes or "")
        print(f"Added: {args.source} -> {args.target}")

    elif args.glossary_action == "remove":
        if glossary.remove(args.source):
            print(f"Removed: {args.source}")
        else:
            print(f"Not found: {args.source}")

    elif args.glossary_action == "list":
        entries = glossary.get_all()
        if not entries:
            print("Glossary is empty.")
            return
        print(f"{'Source':<20} {'Target':<20} {'Notes'}")
        print("-" * 60)
        for e in entries:
            print(f"{e['source']:<20} {e['target']:<20} {e.get('notes', '')}")
        print(f"\nTotal: {len(entries)} entries")

    elif args.glossary_action == "import":
        count = glossary.import_csv(args.csv_file)
        print(f"Imported {count} entries from {args.csv_file}")

    elif args.glossary_action == "export":
        glossary.export_csv(args.csv_file)
        print(f"Exported glossary to {args.csv_file}")


def cmd_stats(args: argparse.Namespace) -> None:
    """Show translation memory statistics."""
    memory = load_memory()
    entries = memory.get("entries", [])
    glossary = Glossary(str(GLOSSARY_PATH))

    print("=== Translation Memory Statistics ===\n")
    print(f"Total entries: {len(entries)}")

    if not entries:
        print("\nNo entries yet. Use 'translate' or 'learn' to add entries.")
        glossary_entries = glossary.get_all()
        print(f"\nGlossary entries: {len(glossary_entries)}")
        return

    # By category
    categories: dict[str, int] = {}
    for e in entries:
        cat = e.get("category", "") or "(uncategorized)"
        categories[cat] = categories.get(cat, 0) + 1

    print("\nBy category:")
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")

    # Auto-approved vs manually approved
    auto = sum(1 for e in entries if e.get("auto_approved", False))
    manual = len(entries) - auto
    print(f"\nAuto-approved: {auto}")
    print(f"Manually approved/corrected: {manual}")

    # Most common translations (by target)
    target_counts: dict[str, int] = {}
    for e in entries:
        tl = e.get("translation", "")
        if tl:
            target_counts[tl] = target_counts.get(tl, 0) + 1

    if target_counts:
        print("\nMost frequently used translations:")
        for tl, count in sorted(target_counts.items(), key=lambda x: -x[1])[:10]:
            preview = tl[:50] + ("..." if len(tl) > 50 else "")
            print(f"  [{count}x] {preview}")

    # Glossary stats
    glossary_entries = glossary.get_all()
    print(f"\nGlossary entries: {len(glossary_entries)}")


def cmd_interactive(args: argparse.Namespace) -> None:
    """Interactive translation mode."""
    translator = _create_translator(
        model_override=args.model if hasattr(args, "model") else None
    )
    glossary = Glossary(str(GLOSSARY_PATH))

    print("=== Interactive Translation Mode ===")
    print(f"Model: {translator._model} (backend: {translator._backend})")
    print("Enter Japanese text to translate. Commands: /quit, /glossary, /stats")
    print()

    while True:
        try:
            text = input("JP> ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nExiting.")
            break

        if not text:
            continue

        if text == "/quit":
            break
        elif text == "/glossary":
            entries = glossary.get_all()
            for e in entries:
                print(f"  {e['source']} -> {e['target']}")
            continue
        elif text == "/stats":
            stats = translator.memory_stats()
            print(f"  Memory entries: {stats['total_entries']}")
            print(f"  Glossary entries: {len(glossary.get_all())}")
            continue

        # Translate using GameTranslator (handles memory lookup + API/CLI call)
        try:
            result = translator.translate(source=text)
            translation = result["translation"]
            confidence = result["confidence"]

            if confidence >= 1.0:
                print(f"EN (from memory)> {translation}")
            else:
                print(f"EN (conf={confidence:.2f})> {translation}")
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            continue

        # Ask for approval
        print("  [a]pprove  [c]orrect  [s]kip")
        try:
            action = input("  > ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            print("\nExiting.")
            break

        if action in ("a", "approve"):
            translator.memory.add(source=text, target=translation)
            print("  Saved to memory.")

        elif action in ("c", "correct"):
            try:
                corrected = input("  Corrected EN> ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nExiting.")
                break
            if corrected:
                translator.memory.add(source=text, target=corrected)
                print("  Correction saved to memory.")

        print()


# ---------------------------------------------------------------------------
# Argument parser
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="ai-translator",
        description="AI Game Translation Tool — ゲーム翻訳AIツール (JP→EN)",
    )
    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # --- translate ---
    p_translate = subparsers.add_parser("translate", help="Translate an Excel file")
    p_translate.add_argument("input", help="Input Excel file path")
    p_translate.add_argument("-o", "--output", help="Output file path (default: input_translated.xlsx)")
    p_translate.add_argument("--category", help="Default category for all entries")
    p_translate.add_argument("--source-col", help="Source text column letter (default: auto-detect)")
    p_translate.add_argument("--threshold", type=float, help="Auto-approve threshold (0.0-1.0)")
    p_translate.add_argument("--model", choices=["opus", "sonnet", "haiku"], help="Override model")
    p_translate.add_argument("--dry-run", action="store_true", help="Show what would be translated without calling API")

    # --- learn ---
    p_learn = subparsers.add_parser("learn", help="Import approved/corrected translations into memory")
    p_learn.add_argument("input", help="Corrected Excel file path")
    p_learn.add_argument("--source-col", help="Source column letter (default: A)")
    p_learn.add_argument("--translation-col", help="Approved translation column letter (default: B)")
    p_learn.add_argument("--corrected-col", help="Corrected translation column letter (if separate)")
    p_learn.add_argument("--category", help="Category for imported entries")

    # --- glossary ---
    p_glossary = subparsers.add_parser("glossary", help="Manage game-specific terminology")
    glossary_sub = p_glossary.add_subparsers(dest="glossary_action", help="Glossary actions")

    p_gl_add = glossary_sub.add_parser("add", help="Add a glossary entry")
    p_gl_add.add_argument("source", help="Source term (Japanese)")
    p_gl_add.add_argument("target", help="Target term (English)")
    p_gl_add.add_argument("--notes", help="Optional notes")

    p_gl_remove = glossary_sub.add_parser("remove", help="Remove a glossary entry")
    p_gl_remove.add_argument("source", help="Source term to remove")

    glossary_sub.add_parser("list", help="List all glossary entries")

    p_gl_import = glossary_sub.add_parser("import", help="Import glossary from CSV")
    p_gl_import.add_argument("csv_file", help="CSV file path (source,target columns)")

    p_gl_export = glossary_sub.add_parser("export", help="Export glossary to CSV")
    p_gl_export.add_argument("csv_file", help="Output CSV file path")

    # --- stats ---
    subparsers.add_parser("stats", help="Show translation memory statistics")

    # --- interactive ---
    p_interactive = subparsers.add_parser("interactive", help="Interactive translation mode")
    p_interactive.add_argument("--model", choices=["opus", "sonnet", "haiku"], help="Override model")

    return parser


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(0)

    try:
        if args.command == "translate":
            cmd_translate(args)
        elif args.command == "learn":
            cmd_learn(args)
        elif args.command == "glossary":
            if not args.glossary_action:
                parser.parse_args(["glossary", "--help"])
            cmd_glossary(args)
        elif args.command == "stats":
            cmd_stats(args)
        elif args.command == "interactive":
            cmd_interactive(args)
        else:
            parser.print_help()
    except KeyboardInterrupt:
        print("\nInterrupted.")
        sys.exit(130)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
