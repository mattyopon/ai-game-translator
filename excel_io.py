"""
Excel I/O module for the game translation tool.

Handles reading source text, writing translation results,
and reading corrected feedback from Excel files.

Requires: openpyxl
"""

from __future__ import annotations

import os
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SOURCE_KEYWORDS = {
    "source", "原文", "japanese", "jp", "日本語", "src", "original",
}
_TRANSLATION_KEYWORDS = {
    "translation", "翻訳", "english", "en", "英語", "trans", "target",
}
_CORRECTED_KEYWORDS = {
    "corrected", "修正", "modified", "correction", "revised", "修正後",
}
_CATEGORY_KEYWORDS = {
    "category", "カテゴリ", "種別", "cat", "type", "分類",
}

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_GREEN_STATUS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_YELLOW_STATUS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _col_to_index(col: str | int) -> int:
    """Convert a column identifier to a 1-based column index.

    Accepts:
      - Column letter(s): "A" -> 1, "B" -> 2, "AA" -> 27
      - Integer (already 1-based): 1, 2, 3, ...
      - Integer as string: "1", "2", "3", ...
    """
    if isinstance(col, int):
        if col < 1:
            raise ValueError(f"Column index must be >= 1, got {col}")
        return col
    col = col.strip()
    if col.isdigit():
        idx = int(col)
        if idx < 1:
            raise ValueError(f"Column index must be >= 1, got {idx}")
        return idx
    try:
        return column_index_from_string(col.upper())
    except (ValueError, AttributeError) as exc:
        raise ValueError(
            f"Invalid column identifier: '{col}'. "
            "Use a column letter (e.g. 'A') or a 1-based number (e.g. 1)."
        ) from exc


def _get_cell_value(ws: Worksheet, row: int, col_idx: int) -> str | None:
    """Read a cell value, handling merged cells gracefully.

    For merged cells, the value is stored in the top-left cell of the merge
    range. This helper walks the merge ranges to find it.
    """
    cell = ws.cell(row=row, column=col_idx)
    value = cell.value

    # If the cell has a value already, return it.
    if value is not None:
        return str(value).strip() if isinstance(value, str) else value

    # Check if this cell is inside a merged range.
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            # The value lives in the top-left cell of the range.
            top_row = merge_range.min_row
            top_col = merge_range.min_col
            top_value = ws.cell(row=top_row, column=top_col).value
            if top_value is not None:
                return str(top_value).strip() if isinstance(top_value, str) else top_value
            return None

    return None


def _open_workbook(path: str, data_only: bool = True) -> openpyxl.Workbook:
    """Open an Excel workbook with helpful error messages."""
    path_obj = Path(path)

    if not path_obj.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    suffix = path_obj.suffix.lower()
    if suffix not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        if suffix == ".xls":
            raise ValueError(
                f"Legacy .xls format is not supported by openpyxl. "
                f"Please convert '{path}' to .xlsx first (e.g. via LibreOffice or Excel)."
            )
        raise ValueError(
            f"Unsupported file extension '{suffix}'. Expected .xlsx (or .xlsm/.xltx/.xltm)."
        )

    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except Exception as exc:
        raise RuntimeError(
            f"Failed to open '{path}': {exc}\n"
            "Make sure the file is a valid Excel (.xlsx) file and is not corrupted."
        ) from exc


def _normalize_header(value: Any) -> str:
    """Normalize a header cell value for keyword matching."""
    if value is None:
        return ""
    return str(value).strip().lower()


def _auto_adjust_column_widths(ws: Worksheet) -> None:
    """Set column widths based on content length."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                text = str(cell.value)
                # Approximate width: CJK characters count as ~2 chars
                length = sum(2 if ord(c) > 0x2E80 else 1 for c in text)
                max_length = max(max_length, length)
        # Clamp between 8 and 60
        adjusted = min(max(max_length + 3, 8), 60)
        ws.column_dimensions[col_letter].width = adjusted


def _detect_first_nonempty_column(ws: Worksheet, start_row: int) -> int:
    """Find the first column (1-based) that has non-empty data starting from *start_row*."""
    for col_idx in range(1, ws.max_column + 1):
        for row_idx in range(start_row, min(start_row + 50, ws.max_row + 1)):
            val = _get_cell_value(ws, row_idx, col_idx)
            if val is not None and str(val).strip():
                return col_idx
    return 1  # fallback


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def detect_columns(path: str, header_row: int = 1) -> dict[str, str | None]:
    """Auto-detect which columns contain what data based on header keywords.

    Scans the header row for known keywords and maps them to roles:
      - Source/原文/Japanese/JP/日本語 -> source column
      - Translation/翻訳/English/EN/英語 -> translation column
      - Corrected/修正/Modified -> corrected column
      - Category/カテゴリ/種別 -> category column

    Parameters
    ----------
    path : str
        Path to the Excel file.
    header_row : int
        Row number (1-based) that contains headers.

    Returns
    -------
    dict
        Mapping of role names to column letters.
        Example: ``{"source": "A", "translation": "B", "corrected": None, "category": "D"}``
    """
    wb = _open_workbook(path)
    ws = wb.active

    result: dict[str, str | None] = {
        "source": None,
        "translation": None,
        "corrected": None,
        "category": None,
    }

    keyword_map: list[tuple[str, set[str]]] = [
        ("source", _SOURCE_KEYWORDS),
        ("translation", _TRANSLATION_KEYWORDS),
        ("corrected", _CORRECTED_KEYWORDS),
        ("category", _CATEGORY_KEYWORDS),
    ]

    for col_idx in range(1, ws.max_column + 1):
        header = _normalize_header(_get_cell_value(ws, header_row, col_idx))
        if not header:
            continue
        for role, keywords in keyword_map:
            if result[role] is not None:
                continue
            # Check both exact match and substring match
            if header in keywords or any(kw in header for kw in keywords):
                result[role] = get_column_letter(col_idx)
                break

    wb.close()
    return result


def read_source_excel(
    path: str,
    source_col: str | int | None = "A",
    category_col: str | int | None = None,
    header_row: int = 1,
    start_row: int = 2,
) -> list[dict]:
    """Read source text from an Excel file.

    Parameters
    ----------
    path : str
        Path to the input Excel file (.xlsx).
    source_col : str | int | None
        Column containing the source text. Accepts a letter ("A") or a
        1-based number (1). If ``None``, auto-detects the first non-empty
        column.
    category_col : str | int | None
        Optional column containing category information.
    header_row : int
        Row number (1-based) of the header row.
    start_row : int
        Row number (1-based) where data starts.

    Returns
    -------
    list[dict]
        Each dict has keys: ``"row"`` (int), ``"source"`` (str), ``"category"`` (str or None).
    """
    wb = _open_workbook(path)
    ws = wb.active

    # Resolve source column
    if source_col is None:
        src_idx = _detect_first_nonempty_column(ws, start_row)
    else:
        src_idx = _col_to_index(source_col)

    # Resolve category column
    cat_idx: int | None = None
    if category_col is not None:
        cat_idx = _col_to_index(category_col)

    results: list[dict] = []

    for row_num in range(start_row, ws.max_row + 1):
        raw_value = _get_cell_value(ws, row_num, src_idx)

        # Skip empty rows
        if raw_value is None:
            continue
        source_text = str(raw_value).strip()
        if not source_text:
            continue

        category: str | None = None
        if cat_idx is not None:
            cat_val = _get_cell_value(ws, row_num, cat_idx)
            if cat_val is not None:
                category = str(cat_val).strip() or None

        results.append({
            "row": row_num,
            "source": source_text,
            "category": category,
        })

    wb.close()
    return results


def write_translation_excel(
    path: str,
    results: list[dict],
    template_path: str | None = None,
    confidence_threshold: float = 0.8,
) -> None:
    """Write translation results to an Excel file.

    Output columns:
      - A: Source (原文)
      - B: Translation (翻訳)
      - C: Confidence (信頼度) — 0.0 to 1.0
      - D: Review Status — ``"AUTO"`` or ``"REVIEW"``
      - E: Similar Examples Used (参考にした類似例の数)
      - F: Category (カテゴリ)
      - G: Notes (備考)

    Formatting applied:
      - Header row: bold white text on blue background.
      - Confidence cells: green (>= 0.8), yellow (0.5–0.8), red (< 0.5).
      - Status cells: green for AUTO, yellow for REVIEW.
      - Column widths auto-adjusted.
      - Freeze panes below the header row.

    Parameters
    ----------
    path : str
        Output file path (.xlsx).
    results : list[dict]
        Each dict should contain at least ``"source"`` and ``"translation"``.
        Optional keys: ``"confidence"`` (float), ``"similar_count"`` (int),
        ``"category"`` (str), ``"notes"`` (str).
    template_path : str | None
        If provided, copies the template and appends translation columns.
    confidence_threshold : float
        Threshold for AUTO vs REVIEW classification (default 0.8).
    """
    if template_path is not None:
        template = Path(template_path)
        if not template.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        shutil.copy2(template_path, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"

    # --- Header row ---
    headers = [
        ("A", "Source (原文)"),
        ("B", "Translation (翻訳)"),
        ("C", "Confidence (信頼度)"),
        ("D", "Review Status"),
        ("E", "Similar Examples (類似例数)"),
        ("F", "Category (カテゴリ)"),
        ("G", "Notes (備考)"),
    ]

    for col_letter, header_text in headers:
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Data rows ---
    for i, entry in enumerate(results, start=2):
        source = entry.get("source", "")
        translation = entry.get("translation", "")
        confidence = entry.get("confidence", 0.0)
        similar_count = entry.get("similar_count", 0)
        category = entry.get("category", "")
        notes = entry.get("notes", "")

        # Determine review status
        if confidence is None:
            confidence = 0.0
        status = "AUTO" if confidence >= confidence_threshold else "REVIEW"

        # A: Source
        cell_a = ws.cell(row=i, column=1, value=source)
        cell_a.border = _THIN_BORDER
        cell_a.alignment = Alignment(vertical="top", wrap_text=True)

        # B: Translation
        cell_b = ws.cell(row=i, column=2, value=translation)
        cell_b.border = _THIN_BORDER
        cell_b.alignment = Alignment(vertical="top", wrap_text=True)

        # C: Confidence
        cell_c = ws.cell(row=i, column=3, value=round(confidence, 3))
        cell_c.border = _THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center", vertical="top")
        cell_c.number_format = "0.00"
        if confidence >= 0.8:
            cell_c.fill = _GREEN_FILL
        elif confidence >= 0.5:
            cell_c.fill = _YELLOW_FILL
        else:
            cell_c.fill = _RED_FILL

        # D: Review Status
        cell_d = ws.cell(row=i, column=4, value=status)
        cell_d.border = _THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center", vertical="top")
        cell_d.font = Font(bold=True)
        if status == "AUTO":
            cell_d.fill = _GREEN_STATUS_FILL
        else:
            cell_d.fill = _YELLOW_STATUS_FILL

        # E: Similar Examples Used
        cell_e = ws.cell(row=i, column=5, value=similar_count if similar_count else 0)
        cell_e.border = _THIN_BORDER
        cell_e.alignment = Alignment(horizontal="center", vertical="top")

        # F: Category
        cell_f = ws.cell(row=i, column=6, value=category or "")
        cell_f.border = _THIN_BORDER
        cell_f.alignment = Alignment(vertical="top", wrap_text=True)

        # G: Notes
        cell_g = ws.cell(row=i, column=7, value=notes or "")
        cell_g.border = _THIN_BORDER
        cell_g.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Formatting ---
    _auto_adjust_column_widths(ws)

    # Freeze panes: freeze below row 1
    ws.freeze_panes = "A2"

    # Auto-filter on the header row
    last_col_letter = get_column_letter(7)
    last_row = max(len(results) + 1, 2)
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Ensure output directory exists
    out_dir = Path(path).parent
    if out_dir != Path(".") and not out_dir.exists():
        out_dir.mkdir(parents=True, exist_ok=True)

    wb.save(path)
    wb.close()


def read_feedback_excel(
    path: str,
    source_col: str | int = "A",
    original_col: str | int = "B",
    corrected_col: str | int | None = None,
    category_col: str | int | None = None,
    header_row: int = 1,
    start_row: int = 2,
) -> list[dict]:
    """Read corrected translations for learning from human feedback.

    Operates in two modes:

    1. **Correction mode** (``corrected_col`` is specified):
       Reads corrections from the given column and returns only entries
       where the corrected text differs from the original translation.

    2. **Approval mode** (``corrected_col`` is ``None``):
       Treats every translation in ``original_col`` as approved.
       Returns all non-empty entries with ``corrected`` equal to ``original``.

    Parameters
    ----------
    path : str
        Path to the feedback Excel file.
    source_col : str | int
        Column containing the source text.
    original_col : str | int
        Column containing the original (machine) translation.
    corrected_col : str | int | None
        Column containing the human-corrected translation. If ``None``,
        approval mode is used.
    category_col : str | int | None
        Optional column containing category information.
    header_row : int
        Row number (1-based) of the header row.
    start_row : int
        Row number (1-based) where data starts.

    Returns
    -------
    list[dict]
        Each dict has keys: ``"source"`` (str), ``"original"`` (str),
        ``"corrected"`` (str), ``"category"`` (str or None).
    """
    wb = _open_workbook(path)
    ws = wb.active

    src_idx = _col_to_index(source_col)
    orig_idx = _col_to_index(original_col)
    corr_idx: int | None = _col_to_index(corrected_col) if corrected_col is not None else None
    cat_idx: int | None = _col_to_index(category_col) if category_col is not None else None

    results: list[dict] = []

    for row_num in range(start_row, ws.max_row + 1):
        source_val = _get_cell_value(ws, row_num, src_idx)
        original_val = _get_cell_value(ws, row_num, orig_idx)

        # Skip rows with no source or no original translation
        if source_val is None or original_val is None:
            continue

        source_text = str(source_val).strip()
        original_text = str(original_val).strip()

        if not source_text or not original_text:
            continue

        # Read category if available
        category: str | None = None
        if cat_idx is not None:
            cat_val = _get_cell_value(ws, row_num, cat_idx)
            if cat_val is not None:
                category = str(cat_val).strip() or None

        if corr_idx is not None:
            # Correction mode: only include rows where a correction was made
            corrected_val = _get_cell_value(ws, row_num, corr_idx)
            if corrected_val is None:
                continue
            corrected_text = str(corrected_val).strip()
            if not corrected_text:
                continue
            # Only include if the corrected text differs from the original
            if corrected_text == original_text:
                continue
            results.append({
                "source": source_text,
                "original": original_text,
                "corrected": corrected_text,
                "category": category,
            })
        else:
            # Approval mode: treat every entry as approved
            results.append({
                "source": source_text,
                "original": original_text,
                "corrected": original_text,
                "category": category,
            })

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Convenience / CLI
# ---------------------------------------------------------------------------

def _summarize_file(path: str) -> None:
    """Print a quick summary of an Excel file (for debugging)."""
    wb = _open_workbook(path)
    ws = wb.active
    print(f"File: {path}")
    print(f"  Sheet: {ws.title}")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
    print(f"  Merged cells: {len(ws.merged_cells.ranges)}")

    # Show detected columns
    cols = detect_columns(path)
    print(f"  Detected columns: {cols}")

    wb.close()


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_io.py <excel_file> [--detect|--read|--summary]")
        print()
        print("Commands:")
        print("  --summary   Show file summary (default)")
        print("  --detect    Detect column roles from headers")
        print("  --read      Read source text and print first 10 entries")
        sys.exit(1)

    filepath = sys.argv[1]
    command = sys.argv[2] if len(sys.argv) > 2 else "--summary"

    if command == "--summary":
        _summarize_file(filepath)
    elif command == "--detect":
        detected = detect_columns(filepath)
        for role, col in detected.items():
            print(f"  {role}: {col or '(not detected)'}")
    elif command == "--read":
        entries = read_source_excel(filepath, source_col=None)
        print(f"Read {len(entries)} entries. First 10:")
        for entry in entries[:10]:
            src = entry["source"]
            if len(src) > 60:
                src = src[:57] + "..."
            cat = entry["category"] or ""
            print(f"  Row {entry['row']:>4}: [{cat}] {src}")
    else:
        print(f"Unknown command: {command}")
        sys.exit(1)
